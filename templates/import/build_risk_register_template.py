"""Builds the Risk Register import template exactly matching what
packages/shared/importing actually parses/validates/persists today.

Column set, required/optional split, enum values, and scoring formulas are
all read directly from:
  - packages/shared/importing/mapping.py (DEFAULT_RISK_REGISTER_MAPPING, DEFERRED_DOMAIN_FIELDS)
  - packages/shared/importing/validation.py (KNOWN_STATUSES, KNOWN_DECISIONS, required fields)
  - packages/shared/importing/transforms.py (parse_int_1_to_5, parse_excel_date, split_cause_event_impact)
  - packages/risk_engine/scoring.py (default_scoring_config: equal weights, band thresholds, reduction formula)

Deliberately EXCLUDES:
  - the 6 "*_calc" reference columns (overall_impact_calc, inherent_score_calc,
    inherent_band_calc, reduction_calc, residual_score_calc, residual_band_calc) —
    the platform always computes these itself; the user asked that calculated
    columns not be visible/fillable.
  - key_controls_ids_or_short_list, actions_link_jira_servicenow_etc, due_date,
    completion, updated_by, last_updated_date — accepted by the column mapper
    but never actually read by row_to_inputs()/create_risk() today, so filling
    them in would silently do nothing. Excluding them avoids that trap.

Regenerate after changing the import mapping/validation/scoring rules:
    source .venv/bin/activate
    python templates/import/build_risk_register_template.py
This overwrites the copy the web app actually serves at
apps/web/public/templates/risk-register-import-template.xlsx — there is
no separate build step, this script writes straight to that path.
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import os

OUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "apps", "web", "public", "templates",
    "risk-register-import-template.xlsx",
)

FONT_NAME = "Arial"

# Each entry: (header, required, kind, comment)
# required: "required" | "conditional" | "optional"
# kind: "text" | "email" | "date" | "scale15" | "enum_status" | "enum_decision" | "long_text"
COLUMNS = [
    (
        "risk_id", "required", "text",
        "A short, unique code for this risk (e.g. RSK-1001). Must not repeat "
        "anywhere else in this sheet — a duplicate blocks the whole import.",
    ),
    (
        "risk_title", "required", "text",
        "A short, human-readable name for the risk. Shown everywhere in the "
        "platform (dashboard, register, reports).",
    ),
    (
        "risk_statement_cause_event_impact", "optional", "long_text",
        "Optional. If you write it in the exact form "
        "'Cause: ... Event: ... Impact: ...' the platform splits it into three "
        "separate fields automatically. Any other text is kept whole as the "
        "risk statement.",
    ),
    (
        "category", "optional", "text",
        "Free text. If it matches an existing category in the platform "
        "exactly (case-sensitive), the risk is filed under it. If it doesn't "
        "match anything, a new category is created automatically with this "
        "name — so a typo creates a stray category rather than failing the "
        "import. Leave blank for uncategorized.",
    ),
    (
        "business_process_value_stream_optional", "optional", "text",
        "Optional. Free text — the business process or value stream this "
        "risk affects.",
    ),
    (
        "team_department", "optional", "text",
        "Optional. Free text — the owning team or department.",
    ),
    (
        "owner_accountable", "optional", "email",
        "The risk owner's email address, EXACTLY matching an existing user "
        "account in the platform. If it doesn't match any account, the risk "
        "is imported with no owner assigned (this does not block the import).",
    ),
    (
        "raised_date", "optional", "date",
        "Optional. Format as a real date (YYYY-MM-DD is safest). When the "
        "risk was first raised.",
    ),
    (
        "next_review_date", "optional", "date",
        "Optional. Format as a real date (YYYY-MM-DD is safest). When this "
        "risk is next due for review.",
    ),
    (
        "status", "optional", "enum_status",
        "One of: draft, open, monitoring, closed. Leave blank or use "
        "anything else and the platform defaults it to 'draft'.",
    ),
    (
        "decision", "optional", "enum_decision",
        "One of: accept, treat, transfer, avoid, pending. Leave blank or use "
        "anything else and the platform defaults it to 'pending'. If you "
        "enter 'accept', 'acceptance_rationale_required_if_accept' becomes "
        "required for that row.",
    ),
    (
        "acceptance_rationale_required_if_accept", "conditional", "long_text",
        "Required ONLY if decision = accept for this row (blocks the import "
        "if missing in that case). Otherwise optional.",
    ),
    (
        "financial_impact_1_5", "required", "scale15",
        "Whole number 1-5. How severe the financial impact would be if this "
        "risk materializes.",
    ),
    (
        "customer_service_impact_1_5", "required", "scale15",
        "Whole number 1-5. Impact on customer service.",
    ),
    (
        "operational_delivery_impact_1_5", "required", "scale15",
        "Whole number 1-5. Impact on operational delivery.",
    ),
    (
        "legal_regulatory_impact_1_5", "required", "scale15",
        "Whole number 1-5. Legal/regulatory impact.",
    ),
    (
        "reputation_impact_1_5", "required", "scale15",
        "Whole number 1-5. Reputational impact.",
    ),
    (
        "health_safety_impact_1_5", "required", "scale15",
        "Whole number 1-5. Health & safety impact.",
    ),
    (
        "likelihood_1_5_12_month_horizon", "required", "scale15",
        "Whole number 1-5. How likely this risk is to occur in the next 12 "
        "months.",
    ),
    (
        "control_effectiveness_1_5", "optional", "scale15",
        "Optional whole number 1-5. How effective the current controls are "
        "at reducing this risk. Leave blank if not yet assessed — the "
        "platform then applies no reduction (see the Scoring tab).",
    ),
    (
        "risk_velocity_optional", "optional", "text",
        "Optional. Free text (e.g. 'fast', 'slow') describing how quickly "
        "impact would be felt once the risk occurs.",
    ),
    (
        "confidence_optional", "optional", "text",
        "Optional. Free text describing your confidence in this assessment.",
    ),
    (
        "treatment_plan_summary", "optional", "long_text",
        "Optional. Free text summary of the treatment plan.",
    ),
    (
        "latest_update_note", "optional", "long_text",
        "Optional. Free text — the most recent status note for this risk.",
    ),
]

EXAMPLE_ROW = {
    "risk_id": "RSK-2001",
    "risk_title": "Single-sourced payment processor outage",
    "risk_statement_cause_event_impact": (
        "Cause: reliance on one payment processor Event: extended processor "
        "outage Impact: inability to take customer payments"
    ),
    "category": "Operational",
    "business_process_value_stream_optional": "Order to cash",
    "team_department": "Finance Operations",
    "owner_accountable": "risk.owner@example.com",
    "raised_date": "2026-01-15",
    "next_review_date": "2026-07-15",
    "status": "open",
    "decision": "treat",
    "acceptance_rationale_required_if_accept": "",
    "financial_impact_1_5": 4,
    "customer_service_impact_1_5": 4,
    "operational_delivery_impact_1_5": 3,
    "legal_regulatory_impact_1_5": 2,
    "reputation_impact_1_5": 3,
    "health_safety_impact_1_5": 1,
    "likelihood_1_5_12_month_horizon": 2,
    "control_effectiveness_1_5": 3,
    "risk_velocity_optional": "fast",
    "confidence_optional": "medium",
    "treatment_plan_summary": "Add a secondary payment processor as failover.",
    "latest_update_note": "Vendor evaluation in progress.",
}

REQUIRED_FILL = PatternFill("solid", fgColor="FBE2E2")
CONDITIONAL_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="E8EEF7")
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10)
EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, color="666666", size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="C9CED6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build():
    wb = Workbook()

    # --- Sheet 1 (MUST be first — the platform only ever reads the first
    # worksheet in the uploaded file, whatever it's named). ---
    ws = wb.active
    ws.title = "Risk Register"

    for col_idx, (header, required, kind, note) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.fill = {
            "required": REQUIRED_FILL,
            "conditional": CONDITIONAL_FILL,
            "optional": OPTIONAL_FILL,
        }[required]
        req_word = {"required": "REQUIRED", "conditional": "CONDITIONALLY REQUIRED", "optional": "optional"}[required]
        cell.comment = Comment(f"{req_word}\n\n{note}", "Risk Intelligence Platform")

        example_val = EXAMPLE_ROW.get(header, "")
        example_cell = ws.cell(row=2, column=col_idx, value=example_val)
        example_cell.font = EXAMPLE_FONT
        example_cell.border = BORDER
        if kind == "date" and example_val:
            example_cell.number_format = "yyyy-mm-dd"

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 26

        # Data validation applied to rows 3-1000 (row 2 is the filled-in
        # example, left as reference — not meant to be deleted, but the
        # validation still doesn't need to cover it).
        rng = f"{col_letter}3:{col_letter}1000"
        if kind == "scale15":
            dv = DataValidation(
                type="whole", operator="between", formula1=1, formula2=5,
                allow_blank=(required != "required"),
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Enter a whole number from 1 to 5.",
            )
            ws.add_data_validation(dv)
            dv.add(rng)
        elif kind == "enum_status":
            dv = DataValidation(
                type="list", formula1='"draft,open,monitoring,closed"',
                allow_blank=True, showErrorMessage=True,
                errorTitle="Invalid status",
                error="Choose one of: draft, open, monitoring, closed.",
            )
            ws.add_data_validation(dv)
            dv.add(rng)
        elif kind == "enum_decision":
            dv = DataValidation(
                type="list", formula1='"accept,treat,transfer,avoid,pending"',
                allow_blank=True, showErrorMessage=True,
                errorTitle="Invalid decision",
                error="Choose one of: accept, treat, transfer, avoid, pending.",
            )
            ws.add_data_validation(dv)
            dv.add(rng)
        elif kind == "date":
            dv = DataValidation(
                type="date", operator="greaterThan", formula1="1900-01-01",
                allow_blank=True, showErrorMessage=True,
                errorTitle="Invalid date",
                error="Enter a real date (e.g. 2026-01-15).",
            )
            ws.add_data_validation(dv)
            dv.add(rng)

        if kind == "date":
            for r in range(3, 1001):
                ws.cell(row=r, column=col_idx).number_format = "yyyy-mm-dd"

    # Block duplicate risk_id within the sheet (blocking error in the real system).
    risk_id_col = get_column_letter(1)
    dup_dv = DataValidation(
        type="custom",
        formula1=f"COUNTIF(${risk_id_col}$3:${risk_id_col}$1000,{risk_id_col}3)<=1",
        allow_blank=True, showErrorMessage=True,
        errorTitle="Duplicate risk_id",
        error="This risk_id is already used elsewhere in this sheet. Each risk_id must be unique.",
    )
    ws.add_data_validation(dup_dv)
    dup_dv.add(f"{risk_id_col}3:{risk_id_col}1000")

    # Highlight required text cells (risk_id, risk_title) if left blank —
    # numeric 1-5 fields already get a hard validation error instead.
    from openpyxl.formatting.rule import FormulaRule

    blank_fill = PatternFill("solid", fgColor="F8D7DA")
    for header_name in ("risk_id", "risk_title"):
        col_idx = [h for h, *_ in COLUMNS].index(header_name) + 1
        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}3:{col_letter}1000",
            FormulaRule(formula=[f'AND(LEN({col_letter}3)=0,COUNTBLANK($A3:$X3)<{len(COLUMNS)})'], fill=blank_fill),
        )

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 60

    # Legend row above the header, in a separate thin strip — can't use
    # row 1 (that's the real header the parser reads), so put the legend on
    # the Instructions sheet instead and just leave a one-line pointer here
    # via the sheet's tab color + a cell comment convention. Simplicity: put
    # a short pointer note in cell A1's neighbor via freeze pane area isn't
    # possible without extra rows, so the Instructions tab carries the full
    # legend; keep this sheet exactly two header/example rows plus data.

    # --- Sheet 2: Instructions & Scoring ---
    ins = wb.create_sheet("Instructions & Scoring")
    ins.sheet_view.showGridLines = False
    ins.column_dimensions["A"].width = 42
    ins.column_dimensions["B"].width = 90

    title_font = Font(name=FONT_NAME, bold=True, size=14)
    h2_font = Font(name=FONT_NAME, bold=True, size=12)
    h3_font = Font(name=FONT_NAME, bold=True, size=10)
    normal_font = Font(name=FONT_NAME, size=10)
    mono_font = Font(name="Consolas", size=10)

    r = 1

    def write(row, col, text, font=normal_font, wrap=True, fill=None):
        c = ins.cell(row=row, column=col, value=text)
        c.font = font
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        if fill:
            c.fill = fill
        return c

    write(r, 1, "Risk Register Import Template", title_font); r += 2

    write(r, 1, "How to use this file", h2_font); r += 1
    write(
        r, 1,
        "1. Fill in the 'Risk Register' tab — one row per risk, starting at row 3 "
        "(row 2 shows a worked example; delete it or leave it and remove it later, "
        "your choice, it will just be validated like any other row).\n"
        "2. Do not rename, reorder, or delete columns, and do not add new columns — "
        "the Import Wizard matches columns by their exact header text.\n"
        "3. Do not reorder the sheet tabs — the platform only reads the FIRST tab "
        "in the workbook as data. This Instructions tab is deliberately second.\n"
        "4. Save as .xlsx and upload it on the Import Wizard page. The wizard shows "
        "you every row it couldn't accept before anything is committed — nothing is "
        "imported silently.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 100
    r += 2

    write(r, 1, "Column color key", h2_font); r += 1
    write(r, 1, "Required — blocks the import if missing", fill=REQUIRED_FILL); write(r, 2, ""); r += 1
    write(r, 1, "Conditionally required — required only in specific cases (see column notes)", fill=CONDITIONAL_FILL); r += 1
    write(r, 1, "Optional", fill=OPTIONAL_FILL); r += 2
    write(
        r, 1,
        "Every column on the Risk Register tab also has a note (hover over the "
        "header cell, or look for the small red triangle) repeating this.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 2

    write(r, 1, "What's NOT in this template, on purpose", h2_font); r += 1
    write(
        r, 1,
        "The platform calculates several fields itself from what you enter — they "
        "are never something you fill in, and are deliberately left out of this "
        "template so there's nothing to get wrong:\n"
        "  • Overall Impact score\n"
        "  • Inherent score and band\n"
        "  • Residual score and band\n"
        "  • The effect of control effectiveness on the residual score\n"
        "See the Scoring section below for exactly how each of these is computed "
        "from the columns you do fill in.\n\n"
        "Controls and Actions (e.g. a list of control IDs, an actions/Jira link, "
        "due dates) are not imported through this Risk Register file in the "
        "current version of the platform — add and link Controls and Actions to "
        "each risk directly in the platform after the import, from that risk's "
        "detail page.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 160
    r += 2

    write(r, 1, "How scoring works (current default configuration)", h2_font); r += 1
    write(
        r, 1,
        "These run automatically after import — you never enter them directly. "
        "An Administrator can change the weights and thresholds below at any time "
        "in the platform's Scoring Config admin page, so treat these as the "
        "defaults, not a permanent guarantee.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 45
    r += 1

    write(r, 1, "1. Overall Impact", h3_font); r += 1
    write(
        r, 1,
        "The average of your six 1-5 impact scores (financial, customer service, "
        "operational delivery, legal/regulatory, reputation, health & safety), "
        "each weighted equally by default: Overall Impact = average of the six "
        "impact scores.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 45
    r += 1

    write(r, 1, "2. Inherent Score & Band", h3_font); r += 1
    write(
        r, 1,
        "Inherent Score = Overall Impact × Likelihood.\n"
        "Band (default thresholds): 0–6 Low · 6–12 Moderate · 12–18 High · 18–25 Extreme.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 30
    r += 1

    write(r, 1, "3. Control Effectiveness → Reduction", h3_font); r += 1
    write(
        r, 1,
        "Reduction = (Control Effectiveness ÷ 5) × 60%. So effectiveness 5 "
        "reduces the inherent score by 60% (the maximum); effectiveness 1 reduces "
        "it by 12%; leaving it blank applies no reduction at all.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 45
    r += 1

    write(r, 1, "4. Residual Score & Band", h3_font); r += 1
    write(
        r, 1,
        "Residual Score = Inherent Score × (1 − Reduction). The same band "
        "thresholds as Inherent (above) are then applied to the residual score to "
        "get the Residual Band — this is the number that drives the heatmap and "
        "most dashboard views.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 45
    r += 2

    write(r, 1, "status values", h2_font); r += 1
    for val, meaning in [
        ("draft", "Not yet finalized / under construction"),
        ("open", "Active and being tracked"),
        ("monitoring", "Being watched, treatment mostly complete"),
        ("closed", "No longer active"),
    ]:
        write(r, 1, val, mono_font); write(r, 2, meaning); r += 1
    write(
        r, 1,
        "Anything else (or blank) is imported as 'draft' — this is a warning, "
        "not a blocking error.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 2

    write(r, 1, "decision values", h2_font); r += 1
    for val, meaning in [
        ("accept", "Risk is knowingly accepted as-is (requires acceptance_rationale_required_if_accept)"),
        ("treat", "Actively being mitigated"),
        ("transfer", "Risk transferred (e.g. insurance, contract)"),
        ("avoid", "Activity causing the risk is being discontinued"),
        ("pending", "Not yet decided"),
    ]:
        write(r, 1, val, mono_font); write(r, 2, meaning); ins.row_dimensions[r].height = 30; r += 1
    write(
        r, 1,
        "Anything else (or blank) is imported as 'pending' — this is a warning, "
        "not a blocking error.",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 2

    write(r, 1, "If the import reports issues", h2_font); r += 1
    write(
        r, 1,
        "The Import Wizard shows every row with a problem before you confirm "
        "anything. Red/blocking issues (missing risk_id or risk_title, an "
        "impact/likelihood score outside 1-5, a duplicate risk_id, or "
        "decision=accept with no rationale) must be fixed in the file and "
        "re-uploaded. Yellow/warning issues (an unrecognized category, owner "
        "email, status, or decision) don't block the import — they just tell "
        "you what the platform did instead (created a new category, left the "
        "risk unowned, defaulted the status/decision).",
    ); ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ins.row_dimensions[r].height = 90

    wb.save(OUT_PATH)
    print("wrote", OUT_PATH)


if __name__ == "__main__":
    build()
