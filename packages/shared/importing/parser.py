"""Reads raw rows/columns out of an uploaded .xlsx file.

Cell values are read as data only — formulas are never evaluated by this
parser (openpyxl with data_only=True reads the last-cached computed value
for a formula cell and never executes one), which forecloses formula-
injection concerns raised in the threat model.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def parse_columns(file_path: str | Path) -> list[str]:
    """Returns the header row (first row) of the first worksheet."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[0]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(cell).strip() if cell is not None else "" for cell in header_row]
    finally:
        wb.close()


def parse_rows(file_path: str | Path) -> list[dict[str, Any]]:
    """Returns every data row (row 2 onward) as {header: value} dicts, in
    file order. A wholly blank row is skipped."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter, ())]
        rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            if all(cell is None or str(cell).strip() == "" for cell in raw_row):
                continue
            row = {header[i]: raw_row[i] for i in range(min(len(header), len(raw_row)))}
            rows.append(row)
        return rows
    finally:
        wb.close()
